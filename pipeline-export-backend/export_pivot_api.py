from fastapi import FastAPI, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

app = FastAPI()

# Allow CORS for local frontend dev
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database connection (PostgreSQL)
DB_USER = os.getenv("DB_USER", "")
DB_PASS = os.getenv("DB_PASS", "")
DB_HOST = os.getenv("DB_HOST", "")
DB_PORT = os.getenv("DB_PORT", "")
DB_NAME = os.getenv("DB_NAME", "")

DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DATABASE_URL)

@app.get("/export-pipeline-template")
def export_pipeline_template():
    query = """
        SELECT id, company_id, deal_name, stage, deal_value, probability, expected_close_date, actual_close_date,
               ae_assigned, sales_manager, lead_source, competitor, loss_reason, notes, last_activity, created_at, updated_at
        FROM deals
    """
    df = pd.read_sql(query, engine)

    # Load your template
    wb = openpyxl.load_workbook('pivot_template.xlsx')
    ws = wb['Deals']

    # Clear existing data (except header)
    ws.delete_rows(2, ws.max_row)

    # Write new data (excluding header)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # Save to output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="pipeline_template.xlsx"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }
    return Response(content=output.read(), headers=headers, media_type=headers["Content-Type"])

@app.get("/export-pipeline-pivot")
def export_pipeline_pivot():
    return export_pipeline_template()
